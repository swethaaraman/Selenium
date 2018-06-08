from selenium import webdriver
import openpyxl
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
# from selenium.webdriver.common.action_chains import ActionChains
import time
path = r"/Users/swethaa/Downloads/chromedriver"
driver = webdriver.Chrome(path)
driver.get("https://finance.yahoo.com/")
listoftickers = ('GOOG', 'MSFT', 'FB', 'PEP', 'AAPL', 'COST')
listofvalues = []
for eachitem in listoftickers:
    driver.find_element_by_name("p").send_keys(eachitem)
    time.sleep(2)
    driver.find_element_by_xpath("""//*[@id="search-button"]""").click()
    time.sleep(2)
    microsoft = driver.find_element_by_xpath("""//*[@id="quote-header-info"]/div[3]/div[1]/div/span[1]""").text
    listofvalues.append(microsoft)
driver.close()
# filling values into excel
fname = 'seleniumresults.xlsx'
workbook = Workbook()
worksheet = workbook.active
Tickers = worksheet.cell(row=1, column=1).value = 'Tickers'
Value = worksheet.cell(row=1, column=2).value = 'Values'
startrow = 2
startcolumn = 1
listoftickers = ('GOOG', 'MSFT', 'FB', 'PEP', 'AAPL', 'COST')
for companies in listoftickers:
    worksheet.cell(row=startrow, column=startcolumn).value = companies
    startrow += 1
beginrow = 2
begincolumn = 2
for values in listofvalues:
    worksheet.cell(row=beginrow, column=begincolumn).value = values
    beginrow += 1
workbook.save(fname)
